from openpyxl import load_workbook
import time
from selenium import webdriver
import locale
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

timeout = 5
lst_row = 0
src_file_name = [
    'IT_Inventory_Printers.xlsx',
]
sheets = ['Принтеры-сканеры-МФУ', ]
wssrc_startrow = 4
typealg_clmn = 'A'
typealg_id = ''
ip_clmn = 'N'
cmn_pause = 2
endrowplus1 = '-1'
ip_notvalid = ['', 'USB', ]
counter_clmn = 'BV'
isprocess_clmn = 'B'
isprocess = 'n'

# 601, 602, 609
tp1url = '/hp/device/InternalPages/Index?id=UsagePage'
tp1xpath = '//*[@id="UsagePage.ImpressionsByMediaSizeTable.Print.A4.Total"]'

# 426
tp2url = '/info_configuration.html?tab=Home&menu=DevConfig'
tp2xpath = '/html/body/div[2]/table/tbody/tr[2]/td[2]/div[7]/table/tbody/tr[1]/td[2]'

# 575
tp3url = tp1url
tp3xpath = '/html/body/div[2]/div/div/div[1]/div/div[2]/div/div/div/div[5]/div/div/div/div/table/tfoot/tr/td[2]'

# 750
tp4url = tp1url
tp4xpath = '/html/body/div[2]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div/div/div/div/table/tbody/tr/td[4]'

# 4525
tp5url = '/hp/device/this.LCDispatcher?nav=hp.Usage'
tp5xpath = '/html/body/table[2]/tbody/tr[2]/td[2]/div[1]/div[5]/table[3]/tbody/tr[2]/td[2]/div'

# 4200
tp6url = '/hp/device/this.LCDispatcher?dispatch=html&cat=0&pos=1'
tp6xpath = '/html/body/div/table[2]/tbody/tr[8]/td[2]/font'

# 630
tp7url = tp1url
tp7xpath = '/html/body/div[2]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div/div/div/div/table/tbody/tr[1]/td[2]'

# 3015
tp8url = tp5url
tp8xpath = '/html/body/table[2]/tbody/tr[2]/td[2]/div[1]/div[3]/table[2]/tbody/tr[3]/td[5]/div'

# 575-2
tp9url = tp1url
tp9xpath = '/html/body/div[2]/div/div/div[2]/div/div[2]/div/div/div/div[3]/div/div/div/div/table/tfoot/tr/td[4]'

# 602-2
tp10url = tp1url
tp10xpath = '/html/body/div[2]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div/div/div/div/table/tbody/tr/td[2]'

# m525
tp11url = '/hp/device/InternalPages/Index?id=UsagePage'
tp11xpath = '/html/body/div[2]/div/div/div[1]/div/div[2]/div/div/div/div[2]/div/div/div/div[1]/table/tbody/tr[2]/td[3]'

chromebtn_advncd = 'Advanced'
chromebtn_prcd = 'Proceed'


def get_last_row(ws):
    lctn = ''
    lst_row = ws.max_row
    max_row = ws.max_row + 1
    print('max_row : ' + str(max_row))
    print('lst_row : ' + str(lst_row))
    for i in range(wssrc_startrow, max_row):
        lctn = ws[typealg_clmn + str(i)].value
        print('typealg : ' + str(lctn))
        if lctn == '' or lctn is None or lctn == endrowplus1:
            lst_row = i - 1
            break
    return lst_row


def prnttext(humaname, element, str_i):
    print('waiting for element:  {}  (sys: {}) for  {}  second(s)'.format(
        humaname, element, str_i))
    print('!!! will exit if more than {} seconds'.format(timeout))


def wfe(driver, element, humaname, byxpath=True, byid=False):
    i = 0
    element_found = False
    while element_found != True:
        return_element = None
        if byxpath:
            try:
                return_element = driver.find_element_by_xpath(element)
                element_found = True
            except:
                element_found = False
                i = i+1
                prnttext(humaname, element, str(i))
                time.sleep(1)
                if i > timeout:
                    print('exit by timeout')
                    break
        if byid:
            try:
                return_element = driver.find_element_by_id(element)
                element_found = True
            except:
                element_found = False
                i = i+1
                prnttext(humaname, element, str(i))
                time.sleep(1)
                if i > timeout:
                    print('exit by timeout')
                    break
    return return_element


def is_float(value):
    try:
        float(value)
        return True
    except:
        return False


def to_float(value):
    if is_float(value):
        return value
    else:
        locale.setlocale(locale.LC_ALL, 'ru_RU')
        return locale.atof(value)


driver = webdriver.Chrome()

for src_file in src_file_name:
    wbsrc = load_workbook(src_file)
    print('opened file {}'.format(src_file))

    for _sheet in sheets:
        wssrc = wbsrc[_sheet]
        print('opened sheet {}'.format(_sheet))

        wssrc_endrow = get_last_row(wssrc)
        #wssrc_endrow = wssrc.max_row

        print('wssrc_endrow {}'.format(str(wssrc_endrow)))

        for row in range(wssrc_startrow, wssrc_endrow + 1, 1):
            print('checking row : ' + str(row))

            typealg_cell = str(typealg_clmn) + str(row)
            ip_cell = str(ip_clmn) + str(row)
            counter_cell = str(counter_clmn) + str(row)

            typealg_id = str(wssrc[typealg_cell].value)
            ip = str(wssrc[ip_cell].value)

            isprocess_cell = str(isprocess_clmn) + str(row)
            isprocess = str(wssrc[isprocess_cell].value)

            if not ip in ip_notvalid and isprocess == 'y':
                try:
                    driver.get('https://{}'.format(ip))

                    counter_ = ''

                    _elm = "//*[contains(text(), '{}')]".format(chromebtn_advncd)
                    inputElement = wfe(driver, _elm, chromebtn_advncd)
                    inputElement.click()

                    _elm = '//*[@id="proceed-link"]'
                    inputElement = wfe(driver, _elm, chromebtn_prcd)
                    inputElement.click()

                    if typealg_id == '1':
                        driver.get("https://{}{}".format(ip, tp1url))
                        _elm = tp1xpath

                    if typealg_id == '2':
                        driver.get("https://{}{}".format(ip, tp2url))
                        _elm = tp2xpath

                    if typealg_id == '3':
                        driver.get("https://{}{}".format(ip, tp3url))
                        _elm = tp3xpath

                    if typealg_id == '4':
                        driver.get("https://{}{}".format(ip, tp4url))
                        _elm = tp4xpath

                    if typealg_id == '5':
                        driver.get("https://{}{}".format(ip, tp5url))
                        _elm = tp5xpath

                    if typealg_id == '6':
                        driver.get("https://{}{}".format(ip, tp6url))
                        _elm = tp6xpath

                    if typealg_id == '7':
                        driver.get("https://{}{}".format(ip, tp7url))
                        _elm = tp7xpath

                    if typealg_id == '8':
                        driver.get("https://{}{}".format(ip, tp8url))
                        _elm = tp8xpath

                    if typealg_id == '9':
                        driver.get("https://{}{}".format(ip, tp9url))
                        _elm = tp9xpath

                    if typealg_id == '10':
                        driver.get("https://{}{}".format(ip, tp10url))
                        _elm = tp10xpath

                    if typealg_id == '11':
                        driver.get("https://{}{}".format(ip, tp11url))
                        _elm = tp11xpath

                    inputElement = wfe(driver, _elm, 'counter')
                    counter_ = inputElement.text

                    counter_ = counter_.replace(',', '')
                    wssrc[counter_cell].value = float(counter_)

                    wbsrc.save(src_file)

                    time.sleep(1)
                except (Exception) as e:
                    print("For row: ", row, " : exception happened : ", e)

    wbsrc.close
driver.quit()
