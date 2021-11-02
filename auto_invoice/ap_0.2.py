import shutil
import openpyxl
import num2t4ru
from num2t4ru import num2text, decimal2text
from decimal import *
from openpyxl.styles.borders import Border, Side

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')

rcv_file_folder = r'C:\...\data_out'

src_file_name = r'C:\...\data_in\data1.xlsx'

tmpl_file_name = r'C:\...\template\template1.xlsx'

sheet = 'Лист1'

rcv_file_name_templ = '{}.xlsx'
rcv_file_name = ''

wbsrc = openpyxl.load_workbook(src_file_name)
wssrc = wbsrc.get_sheet_by_name(sheet)

company_name_clmn = 'B'
company_inn_clmn = 'C'
date_clmn = 'E'
num_nakl_clmn = 'D'
sum_nakl_clmn = 'F'
dolg_clmn = 'G'

company_name = ''  # B
company_name_new = ''
company_inn = ''  # C
date = ''  # E
num_nakl = ''  # D
sum_nakl = 0  # F
dolg = 0  # G

wssrc_startrow = 2
wssrc_endrow = wssrc.max_row

rcv_company_name_cell = 'H10'
rcv_company_inn_cell = 'H11'
rcv_summ_prop_cell = 'D16'
rcv_summ_decml_cell = 'B16'

rcv_row_num_clmn = 'B'
rcv_date_clmn = 'D'
rcv_num_nakl_clmn = 'E'
rcv_sum_nakl_clmn = 'F'
rcv_dolg_clmn = 'G'

row_ins = 19  # перед ней делается инсерт строки в первый раз
row_inc = 0  # сдвиг строки если несколько вставлять

rcv_file_name_templ_comp = ''

sum_nakl_aggregate = Decimal(0)
sum_dolg_aggregate = Decimal(0)

sum_nakl_clmn = 'F'
sum_dolg_clmn = 'G'

for row_num in range(wssrc_startrow, wssrc_endrow + 1):
    company_name = wssrc['{}{}'.format(company_name_clmn, row_num)].value
    company_name = company_name.replace('"', ' ')
    company_name = company_name.replace('\'', ' ')
    company_name = company_name.strip()

    # 1й вход или изменение должника
    if company_name_new != company_name:
        row_inc = 0

        if company_name_new == '':  # только начали, 1ая запись
            company_name_new = company_name
        else:  # не 1й вход, надо сохранить и закрыть файл
            wsrcv[str(sum_nakl_clmn) + str(rcv_row+1)
                  ].value = sum_nakl_aggregate
            wsrcv[str(sum_dolg_clmn) + str(rcv_row+1)
                  ].value = sum_dolg_aggregate

            wsrcv[rcv_summ_prop_cell].value = decimal2text(
                Decimal(sum_dolg_aggregate), int_units=int_units, exp_units=exp_units) + ' )'
            wsrcv[rcv_summ_decml_cell].value = Decimal(sum_dolg_aggregate)

            # сумма в низ
            wsrcv['D'+str(rcv_row+4)].value = Decimal(sum_dolg_aggregate)

            sum_nakl_aggregate = 0
            sum_dolg_aggregate = 0

            wbrcv.save(rcv_file_name)
            wbrcv.close()
            print('closed ', rcv_file_name_templ_comp)
            company_name_new = company_name

        rcv_file_name_templ_comp = rcv_file_name_templ.format(company_name)
        rcv_file_name = rcv_file_folder + '\\' + rcv_file_name_templ_comp
        shutil.copy(tmpl_file_name, rcv_file_name)
        wbrcv = openpyxl.load_workbook(rcv_file_name)
        wsrcv = wbrcv.get_sheet_by_name(sheet)

        company_inn = str(
            wssrc['{}{}'.format(company_inn_clmn, row_num)].value)
        date = str(wssrc['{}{}'.format(date_clmn, row_num)].value)
        date = date[:10]
        num_nakl = str(wssrc['{}{}'.format(num_nakl_clmn, row_num)].value)
        sum_nakl = Decimal(wssrc['{}{}'.format(sum_nakl_clmn, row_num)].value)
        dolg = Decimal(wssrc['{}{}'.format(dolg_clmn, row_num)].value)

        sum_nakl_aggregate += Decimal(sum_nakl)
        sum_dolg_aggregate += Decimal(dolg)

        # вписываем имя компании и инн
        wsrcv[rcv_company_name_cell].value = company_name
        wsrcv[rcv_company_inn_cell].value = company_inn

        # вставляем строку к заполнению
        rcv_row = row_ins + row_inc
        wsrcv.insert_rows(rcv_row, 1)
        print('added row to ', rcv_file_name_templ_comp)
        row_inc += 1

        # заполняем вставленную строку
        # № по порядку
        wsrcv[str(rcv_row_num_clmn) + str(rcv_row)].value = row_inc

        # data
        wsrcv[str(rcv_date_clmn) + str(rcv_row)].value = date

        # nomer nakl
        wsrcv[str(rcv_num_nakl_clmn) + str(rcv_row)].value = num_nakl

        # summa nakl
        wsrcv[str(rcv_sum_nakl_clmn) + str(rcv_row)].value = sum_nakl

        # summa dolga
        wsrcv[str(rcv_dolg_clmn) + str(rcv_row)].value = dolg

        # borders
        wsrcv.cell(row=rcv_row, column=2).border = thin_border
        wsrcv.cell(row=rcv_row, column=3).border = thin_border
        wsrcv.cell(row=rcv_row, column=4).border = thin_border
        wsrcv.cell(row=rcv_row, column=5).border = thin_border
        wsrcv.cell(row=rcv_row, column=6).border = thin_border
        wsrcv.cell(row=rcv_row, column=7).border = thin_border

    # не 1й вход, уже читали минимум 1 раз, следующая строка по той же компании
    elif company_name_new == company_name:
        date = str(wssrc['{}{}'.format(date_clmn, row_num)].value)
        date = date[:10]
        num_nakl = str(wssrc['{}{}'.format(num_nakl_clmn, row_num)].value)
        sum_nakl = Decimal(wssrc['{}{}'.format(sum_nakl_clmn, row_num)].value)
        dolg = Decimal(wssrc['{}{}'.format(dolg_clmn, row_num)].value)

        sum_nakl_aggregate += Decimal(sum_nakl)
        sum_dolg_aggregate += Decimal(dolg)

        # вставляем строку к заполнению
        rcv_row = row_ins + row_inc
        wsrcv.insert_rows(rcv_row, 1)
        print('added row to ', rcv_file_name_templ_comp)
        row_inc += 1

        # заполняем вставленную строку
        # № по порядку
        wsrcv[str(rcv_row_num_clmn) + str(rcv_row)].value = row_inc

        # data
        wsrcv[str(rcv_date_clmn) + str(rcv_row)].value = date

        # nomer nakl
        wsrcv[str(rcv_num_nakl_clmn) + str(rcv_row)].value = num_nakl

        # summa nakl
        wsrcv[str(rcv_sum_nakl_clmn) + str(rcv_row)].value = sum_nakl

        # summa dolga
        wsrcv[str(rcv_dolg_clmn) + str(rcv_row)].value = dolg

        # borders
        wsrcv.cell(row=rcv_row, column=2).border = thin_border
        wsrcv.cell(row=rcv_row, column=3).border = thin_border
        wsrcv.cell(row=rcv_row, column=4).border = thin_border
        wsrcv.cell(row=rcv_row, column=5).border = thin_border
        wsrcv.cell(row=rcv_row, column=6).border = thin_border
        wsrcv.cell(row=rcv_row, column=7).border = thin_border

    wbrcv.save(rcv_file_name)
