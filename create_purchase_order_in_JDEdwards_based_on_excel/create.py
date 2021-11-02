from openpyxl import load_workbook
import time
from selenium import webdriver
import SECURE_DATA
import locale
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

lst_row = 0


def get_last_row(ws):
    lctn = ''
    lst_row = ws.max_row
    max_row = ws.max_row + 1
    print('max_row : ' + str(max_row))
    print('lst_row : ' + str(lst_row))
    for i in range(1, max_row):
        lctn = ws['B' + str(i)].value
        print('lctn : ' + str(lctn))
        if lctn == '' or lctn is None:
            lst_row = i - 1
            break
    return lst_row


src_file_name = [
    'bills_registration.xlsx',
]

sheets = ['Accruals', ]

supp_jde_code_clm = 'D'
cc_code_clm = 'E'
year_clm = 'F'
month_clm = 'G'
summ_clm = 'H'
purpose_clm = 'B'
po_num_column = 'I'
is_load_column = 'L'
vatType_clm = 'M'
Quantity_clm = 'O'

curr_column = 'J'
CURRS = ['RUB', 'USD']

is_load_yes = 'x'
cmn_pause = 2

########

# Известная проблема:

########


def wfe(driver, element, humaname, byxpath=True, byid=False):
    i = 0
    element_found = False
    while element_found != True:
        return_element = None
        if byxpath:
            try:
                return_element = driver.find_element_by_xpath(element)
                element_found = True
            except:
                element_found = False
                i = i+1
                print('waiting for element:  {}  (sys: {}) for  {}  second(s)'.format(
                    humaname, element, str(i)))
                time.sleep(1)
        if byid:
            try:
                return_element = driver.find_element_by_id(element)
                element_found = True
            except:
                element_found = False
                i = i+1
                print('waiting for element:  {}  (sys: {}) for  {}  second(s)'.format(
                    humaname, element, str(i)))
                time.sleep(1)
    return return_element


def is_float(value):
    try:
        float(value)
        return True
    except:
        return False


def to_float(value):
    if is_float(value):
        return value
    else:
        locale.setlocale(locale.LC_ALL, 'ru_RU')
        return locale.atof(value)


driver = webdriver.Chrome()
driver.get("http://jde-web-pd.corpdomen.local:7080/jde/")

_elm = '//*[@id="User"]'
inputElement = wfe(driver, _elm, 'поле логина')
inputElement.click()
inputElement.send_keys(SECURE_DATA.login)

# pass
_elm = 'Password'
inputElement = wfe(driver, _elm, 'поле пароля', False, True)
inputElement.click()
inputElement.send_keys(SECURE_DATA.password)

# btn sign in
_elm = '//*[@id="mainLoginTable"]/tbody/tr[7]/td/input'
inputElement = wfe(driver, _elm, 'кнопка логина')
inputElement.click()

print('logged in')

for src_file in src_file_name:
    wbsrc = load_workbook(src_file)
    print('opened file')

    for _sheet in sheets:
        wssrc = wbsrc[_sheet]
        print('opened sheet')

        wssrc_startrow = 1
        wssrc_endrow = wssrc.max_row

        print('wssrc_endrow : ' + str(wssrc_endrow))

        for row in range(wssrc_startrow, wssrc_endrow + 1, 1):
            print('checking is need to load row : ' + str(row))

            purpose_clm_cell = str(purpose_clm) + str(row)
            supp_jde_code_clm_cell = str(supp_jde_code_clm) + str(row)
            cc_code_clm_cell = str(cc_code_clm) + str(row)
            year_clm_cell = str(year_clm) + str(row)
            month_clm_cell = str(month_clm) + str(row)
            summ_clm_cell = str(summ_clm) + str(row)
            po_num_column_cell = str(po_num_column) + str(row)
            is_load_column_cell = str(is_load_column) + str(row)
            curr_cell = str(curr_column) + str(row)
            vatType_column_cell = str(vatType_clm) + str(row)
            Quantity_cell = str(Quantity_clm) + str(row)

            is_load = str(wssrc[is_load_column_cell].value)

            if is_load == is_load_yes:
                print('грузим строку : ' + str(row))
                # читаем строку из файла
                purpose = str(wssrc[purpose_clm_cell].value)
                supp_jde_code = str(wssrc[supp_jde_code_clm_cell].value)
                cc_code = str(wssrc[cc_code_clm_cell].value)
                year = str(wssrc[year_clm_cell].value)
                month = str(wssrc[month_clm_cell].value)
                vatType = str(wssrc[vatType_column_cell].value)
                Quantity = str(wssrc[Quantity_cell].value)

                curr = str(wssrc[curr_cell].value)
                if curr not in CURRS:
                    curr = 'RUB'

                summ_flt = to_float(wssrc[summ_clm_cell].value)
                summ = str(summ_flt)
                summ = summ.replace('.', ',')

                purpose_text = '{}'.format(purpose)
                # purpose_text = '{} {}-{}'.format(purpose, month, year)

                # открываем Enter Non-Stock Purchase Orders
                driver.get('http://jde-web-pd.corpdomen.local:7080/jde/ShortcutLauncher?OID=P4310_W4310G_ZH010&FormDSTmpl=|1|2|3|6|7|8|9|10|11|12|13|14|15|16|17|20|&FormDSData=|||||||||||||||||')

                # нажимаем +
                _elm = '//*[@id="outer0_76"]'
                inputElement = wfe(driver, _elm, 'нажимаем +')
                inputElement.click()

                # вводим supplier
                _elm = '//*[@id="C0_715"]'
                _type = 'supplier'
                _val = supp_jde_code
                _pause = 0
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                inputElement.send_keys(_val)
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # вводим Currency
                _elm = '//*[@id="C0_20"]'
                _type = 'currency'
                _val = curr
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                inputElement.send_keys(_val)
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # вводим CC code
                _elm = '//*[@id="G0_1_R0"]/td[4]/div/input'
                _type = 'CCcode'
                _val = cc_code
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))
                inputElement.send_keys(_val)

                # Unit_cost
                _elm = '//*[@id="G0_1_R0"]/td[9]/div/input'
                _type = 'Unit_cost'
                _val = (str(summ)).replace('.', ',')
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))
                inputElement.send_keys(_val)

                # клик в Quantity Ordered
                _elm = '//*[@id="G0_1_R0"]/td[5]/div/input'
                _type = 'Quantity Ordered'

                _val = Quantity
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))
                inputElement.send_keys(_val)

                # клик в Description_1
                _elm = '//*[@id="G0_1_R0"]/td[13]/div/input'
                _type = 'Description_1'
                _val = purpose_text
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))
                inputElement.send_keys(_val)

                # клик в input выпадающего Tax Y/N
                _type = 'Tax Y/N'
                if (vatType == is_load_yes):
                    _val = 8  # Стоит x => Поставщик работает без НДС, сумма в РО без НДС
                else:
                    _val = 6  # НДС 20%, сумма в РО без НДС
                # делаем -- ТАБов чтобы прокрутить гор. линию прокрутки
                N = 22  # number of times you want to press TAB
                actions = ActionChains(driver)
                actions.send_keys(Keys.TAB * N)
                actions.perform()

                # ищем номер колонки с текстом Tax\nY/N (убрал T чтоб find вернул 1)
                val = 0
                for clmnind in range(1, 100):
                    try:
                        inputElement = driver.find_element_by_xpath(
                            '//td[@colindex="{}"]'.format(clmnind))
                        print(inputElement.text)
                        print(clmnind)
                        if ((inputElement.text.find('ax') == 1) and (inputElement.text.find('Y/N') > 0)):
                            val = clmnind
                            break
                    except:
                        yes_except = 1

                # ищем все элементы с номером колонки и среди них - тот у которого есть дочерний инпут
                inputElements = driver.find_elements_by_xpath(
                    '//td[@colindex="{}"]'.format(val))
                for elem in inputElements:
                    try:
                        elem2 = elem.find_element_by_css_selector('input')
                        elem2.click()
                        elem2.send_keys(_val)
                    except:
                        yes_except = 1

                _pause = 1
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # считываем РО
                _elm = '//*[@id="C0_22"]'
                _type = 'PO_num'
                inputElement = wfe(driver, _elm, 'ищем {}'.format(_type))
                po_num = inputElement.get_attribute('value')

                # пишем PO в файл
                wssrc[po_num_column_cell].value = po_num

                # убираем X в файле
                wssrc[str('L') + str(row)].value = ''

                # клик на галку
                _elm = '//*[@id="hc_OK"]'
                _type = 'Click_OK'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, '{}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # Close window
                _elm = '//*[@id="jdeclose_ena"]'
                _type = 'Click_Close'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, '{}'.format(_type))
                #inputElement = driver.find_element_by_xpath(_elm)
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # сохраняем файл
                wbsrc.save(src_file)  # src_file_name

                # открываем печать ##################################################################################################
                driver.get('http://jde-web-pd.corpdomen.local:7080/jde/ShortcutLauncher?OID=P98305W_W98305WD&FormDSTmpl=|1|2|3|4|5|6|7|8|9|11|12|&FormDSData=|R5743500|ZH032||||03.09.20|4|0||||')
                for i in range(1, 1):
                    time.sleep(1)
                    print('sleep waiting Printing window ' + str(i))

                # data selection
                _elm = '//*[@id="C0_23"]'
                _type = 'Data_select'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))
                # inputElement.send_keys(_val)

                # submit
                _elm = '//*[@id="divC0_30"]'
                _type = 'Submit'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))
                # inputElement.send_keys(_val)

                # dropdown with type
                _elm = "//select[@name='RightOperand6']/option[text()='Literal']"
                _type = 'Select = Literal'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                #inputElement = driver.find_element_by_xpath(_elm)
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # input PO
                _elm = '//*[@id="LITtf"]'
                _type = 'input PO num'
                _val = po_num
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.send_keys(_val)
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))


                # OK
                _elm = '//*[@id="hc_Select"]'
                _type = 'OK'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))


                # OK common form
                _elm = '//*[@id="hc_Select"]'
                _type = 'OK'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

                # OK printer form
                _elm = '//*[@id="hc_OK"]'
                _type = 'Sent to Print PDF'
                _val = _type
                _pause = 1
                inputElement = wfe(driver, _elm, 'вводим {}'.format(_type))
                inputElement.click()
                for i in range(1, cmn_pause):
                    time.sleep(1)
                    print('{} input: {} - time {}'.format(_type, _val, str(i)))

    wbsrc.close
driver.quit()
